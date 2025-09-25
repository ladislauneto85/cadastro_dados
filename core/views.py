from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.contrib import messages
from datetime import datetime
from django.db.models import Case, When, Value, IntegerField
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font

from pessoal.models import InformacoesPessoais
from funcional.models import InformacoesFuncionais, Documento
from familiar.models import InformacoesFamiliares, Filho
from pessoal.forms import InformacoesPessoaisForm
from funcional.forms import InformacoesFuncionaisForm
from familiar.forms import InformacoesFamiliaresForm

def is_admin(user):
    return user.is_superuser

def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('admin_visualizacao')
        else:
            return redirect('perfil_usuario')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_superuser:
                    return redirect('admin_visualizacao')
                else:
                    return redirect('perfil_usuario')
        else:
            messages.error(request, "Usuário ou senha inválidos.")
    else:
        form = AuthenticationForm()
    return render(request, 'core/login.html', {'form': form})


@login_required
@user_passes_test(is_admin)
def admin_visualizacao(request):
    posto_order = Case(
        When(info_funcionais__posto_graduacao='coronel', then=Value(1)),
        When(info_funcionais__posto_graduacao='tenente_coronel', then=Value(2)),
        When(info_funcionais__posto_graduacao='major', then=Value(3)),
        When(info_funcionais__posto_graduacao='capitao', then=Value(4)),
        When(info_funcionais__posto_graduacao='tenente', then=Value(5)),
        When(info_funcionais__posto_graduacao='aspirante', then=Value(6)),       
        When(info_funcionais__posto_graduacao='subtenente', then=Value(7)),      
        When(info_funcionais__posto_graduacao='sargento', then=Value(8)),        
        When(info_funcionais__posto_graduacao='cabo', then=Value(9)),            
        When(info_funcionais__posto_graduacao='soldado', then=Value(10)),        
        default=Value(11),                                                       
        output_field=IntegerField(),
    )
    
    usuarios = User.objects.filter(is_superuser=False).select_related(
        'info_pessoais', 'info_funcionais'
    ).annotate(
        posto_order=posto_order
    ).order_by('posto_order', 'info_funcionais__matricula')

    return render(request, 'core/admin_visualizacao.html', {'usuarios': usuarios})


@login_required
@user_passes_test(is_admin)
def detalhe_usuario(request, user_id):
    # Primeiro, buscamos o usuário específico que o admin quer ver.
    # Se não existir, a página retornará um erro 404 (Página Não Encontrada).
    usuario_selecionado = get_object_or_404(User, id=user_id)

    # Agora, buscamos cada bloco de informação relacionado a esse usuário.
    # Usamos try/except para o caso de o usuário ainda não ter preenchido aquela seção.
    
    info_pessoais = None
    try:
        # Para OneToOneField, podemos acessar diretamente.
        info_pessoais = usuario_selecionado.info_pessoais
    except InformacoesPessoais.DoesNotExist:
        pass  # Se não existir, a variável continua como None

    info_funcionais = None
    try:
        # Usamos .prefetch_related() para otimizar a busca dos documentos.
        info_funcionais = InformacoesFuncionais.objects.prefetch_related('documentos').get(usuario=usuario_selecionado)
    except InformacoesFuncionais.DoesNotExist:
        pass

    info_familiares = None
    try:
        # Otimizamos também a busca dos filhos.
        info_familiares = InformacoesFamiliares.objects.prefetch_related('filhos').get(usuario=usuario_selecionado)
    except InformacoesFamiliares.DoesNotExist:
        pass
    
    # Montamos o contexto para enviar ao template.
    context = {
        'usuario_selecionado': usuario_selecionado,
        'info_pessoais': info_pessoais,
        'info_funcionais': info_funcionais,
        'info_familiares': info_familiares,
    }
    
    return render(request, 'core/detalhe_usuario.html', context)


# Em core/views.py

@login_required
@transaction.atomic
def perfil_usuario_view(request):
    """
    View para que o usuário comum edite suas informações, agora com
    lógica não-destrutiva completa para adicionar e remover documentos e filhos.
    """
    if request.user.is_superuser:
        return redirect('admin_visualizacao')

    user = request.user
    pessoal_info, _ = InformacoesPessoais.objects.get_or_create(usuario=user)
    funcional_info, _ = InformacoesFuncionais.objects.get_or_create(usuario=user)
    familiar_info, _ = InformacoesFamiliares.objects.get_or_create(usuario=user)

    if request.method == 'POST':
        pessoal_form = InformacoesPessoaisForm(request.POST, request.FILES, instance=pessoal_info)
        funcional_form = InformacoesFuncionaisForm(request.POST, instance=funcional_info)
        familiar_form = InformacoesFamiliaresForm(request.POST, instance=familiar_info)

        if pessoal_form.is_valid() and funcional_form.is_valid() and familiar_form.is_valid():
            pessoal_form.save()
            funcional_instance = funcional_form.save()
            familiar_instance = familiar_form.save()

            # --- LÓGICA DE DOCUMENTOS NÃO-DESTRUTIVA ---
            # 1. Processa as exclusões solicitadas
            documentos_ids_para_deletar = request.POST.getlist('documentos_a_deletar')
            if documentos_ids_para_deletar:
                Documento.objects.filter(id__in=documentos_ids_para_deletar, info_funcional=funcional_instance).delete()

            # 2. Processa as novas adições
            documentos_nomes = request.POST.getlist('nome_documento')
            documentos_arquivos = request.FILES.getlist('arquivo_documento')
            for i, nome in enumerate(documentos_nomes):
                if nome and i < len(documentos_arquivos):
                    Documento.objects.create(
                        info_funcional=funcional_instance,
                        nome_documento=nome,
                        arquivo=documentos_arquivos[i]
                    )

            # --- LÓGICA DE FILHOS NÃO-DESTRUTIVA (AGORA COMPLETA) ---
            # 1. Processa as exclusões solicitadas
            filhos_ids_para_deletar = request.POST.getlist('filhos_a_deletar')
            if filhos_ids_para_deletar:
                Filho.objects.filter(id__in=filhos_ids_para_deletar, info_familiar=familiar_instance).delete()

            # 2. Processa as novas adições
            filhos_nomes_novos = request.POST.getlist('nome_filho_novo')
            filhos_nascimentos_novos = request.POST.getlist('nascimento_filho_novo')
            for i, nome in enumerate(filhos_nomes_novos):
                if nome and i < len(filhos_nascimentos_novos) and filhos_nascimentos_novos[i]:
                    data_nasc_obj = datetime.strptime(filhos_nascimentos_novos[i], '%Y-%m-%d').date()
                    Filho.objects.create(
                        info_familiar=familiar_instance,
                        nome=nome,
                        data_nascimento=data_nasc_obj
                    )

            messages.success(request, 'Suas informações foram salvas com sucesso!')
            return redirect('perfil_usuario')
        else:
            # Lógica de depuração para ver erros de validação
            print("--- ERROS DE VALIDAÇÃO ---")
            if pessoal_form.errors: print("Erros Pessoais:", pessoal_form.errors.as_json())
            if funcional_form.errors: print("Erros Funcionais:", funcional_form.errors.as_json())
            if familiar_form.errors: print("Erros Familiares:", familiar_form.errors.as_json())
            print("--------------------------")
            messages.error(request, 'Houve um erro no formulário. Por favor, verifique os dados.')

    # O resto da view permanece igual
    pessoal_form = InformacoesPessoaisForm(instance=pessoal_info)
    funcional_form = InformacoesFuncionaisForm(instance=funcional_info)
    familiar_form = InformacoesFamiliaresForm(instance=familiar_info)
    documentos_existentes = funcional_info.documentos.all()
    filhos_existentes = familiar_info.filhos.all()

    context = {
        'pessoal_form': pessoal_form,
        'funcional_form': funcional_form,
        'familiar_form': familiar_form,
        'documentos_existentes': documentos_existentes,
        'filhos_existentes': filhos_existentes,
    }
    return render(request, 'core/perfil_usuario.html', context)


@login_required
@user_passes_test(is_admin)
def exportar_excel_view(request):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)  # Remove a planilha padrão

    # Otimiza a busca no banco de dados
    usuarios = User.objects.filter(is_superuser=False).select_related(
        'info_pessoais', 'info_funcionais', 'info_familiares'
    ).prefetch_related(
        'info_funcionais__documentos', 'info_familiares__filhos'
    )

    for usuario in usuarios:
        # Garante que o nome da planilha seja válido
        sheet_title = ''.join(filter(str.isalnum, usuario.username))[:31]
        sheet = workbook.create_sheet(title=sheet_title)
        
        header_font = Font(bold=True)
        sheet['A1'] = "Seção"
        sheet['B1'] = "Campo"
        sheet['C1'] = "Valor"
        sheet['A1'].font = header_font
        sheet['B1'].font = header_font
        sheet['C1'].font = header_font

        row = 2
        def add_data(secao, campo, valor):
            nonlocal row
            sheet.cell(row=row, column=1, value=secao)
            sheet.cell(row=row, column=2, value=campo)
            # Formata datas para um formato legível, se for um objeto date/datetime
            if hasattr(valor, 'strftime'):
                valor = valor.strftime('%d/%m/%Y')
            sheet.cell(row=row, column=3, value=str(valor) if valor is not None else "")
            row += 1

        # --- Informações Pessoais ---
        try:
            p = usuario.info_pessoais
            add_data("Pessoal", "Nome Completo", p.nome_completo)
            add_data("Pessoal", "CPF", p.cpf)
            add_data("Pessoal", "RG", p.rg)
            add_data("Pessoal", "Email", p.email)
            add_data("Pessoal", "Data de Nascimento", p.data_nascimento)
            add_data("Pessoal", "Idade", p.idade)
            add_data("Pessoal", "Telefone", p.telefone)
            add_data("Pessoal", "Grupo Sanguíneo", p.grupo_sanguineo)
            add_data("Pessoal", "Estado Civil", p.get_estado_civil_display())
            add_data("Pessoal", "Escolaridade", p.get_escolaridade_display())
            add_data("Pessoal", "Nº CNH", p.cnh_numero)
            add_data("Pessoal", "Categoria CNH", p.cnh_categoria)
            add_data("Pessoal", "Validade CNH", p.cnh_validade)
            add_data("Pessoal", "Título de Eleitor", p.titulo_eleitor)
            add_data("Pessoal", "Zona Eleitoral", p.zona)
            add_data("Pessoal", "Seção Eleitoral", p.secao)
            add_data("Pessoal", "Município de Votação", p.municipio_votacao)
        except InformacoesPessoais.DoesNotExist:
            add_data("Pessoal", "Status", "Não cadastrado")

        # --- Informações Funcionais ---
        try:
            f = usuario.info_funcionais
            add_data("Funcional", "Nome de Guerra", f.nome_guerra)
            add_data("Funcional", "Posto/Graduação", f.get_posto_graduacao_display())
            add_data("Funcional", "Matrícula", f.matricula)
            add_data("Funcional", "Data de Admissão", f.data_admissao)
            add_data("Funcional", "Banco", f.banco)
            add_data("Funcional", "Agência", f.agencia)
            add_data("Funcional", "Conta Corrente", f.conta_corrente)
        except InformacoesFuncionais.DoesNotExist:
            add_data("Funcional", "Status", "Não cadastrado")

        # --- Informações Familiares ---
        try:
            fam = usuario.info_familiares
            add_data("Familiar", "Endereço", f"{fam.endereco}, {fam.numero_casa} - {fam.bairro}")
            add_data("Familiar", "Complemento", fam.complemento)
            add_data("Familiar", "Cidade", fam.cidade)
            add_data("Familiar", "CEP", fam.cep)
            add_data("Familiar", "Nome do Cônjuge", fam.conjuge_nome)
            add_data("Familiar", "Data Nasc. Cônjuge", fam.conjuge_data_nascimento)
            
            # --- Filhos ---
            if fam.filhos.exists():
                for i, filho in enumerate(fam.filhos.all()):
                    add_data("Filhos", f"Filho(a) {i+1} - Nome", filho.nome)
                    add_data("Filhos", f"Filho(a) {i+1} - Data Nasc.", filho.data_nascimento)
            else:
                add_data("Filhos", "Status", "Nenhum filho cadastrado")

        except InformacoesFamiliares.DoesNotExist:
            add_data("Familiar", "Status", "Não cadastrado")
            
        # Ajustar largura das colunas
        sheet.column_dimensions['A'].width = 15
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 40

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="cadastros_usuarios.xlsx"'
    workbook.save(response)
    return response


@login_required
@user_passes_test(is_admin)
@transaction.atomic
def cadastro_admin_view(request):
    # Esta view é opcional, para o admin criar um usuário completo de uma vez
    if request.method == 'POST':
        # ... (lógica da view cadastro_usuario anterior, se desejar mantê-la)
        return redirect('admin_visualizacao')
    
    pessoal_form = InformacoesPessoaisForm()
    funcional_form = InformacoesFuncionaisForm()
    familiar_form = InformacoesFamiliaresForm()
    context = {
        'pessoal_form': pessoal_form,
        'funcional_form': funcional_form,
        'familiar_form': familiar_form
    }
    return render(request, 'core/cadastro_admin.html', context)